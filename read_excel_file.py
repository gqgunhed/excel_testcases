#!/usr/bin/env python

from openpyxl import Workbook, load_workbook
wb = Workbook()

# open existing file
wb = load_workbook('test1.xlsx')

# print contents
for row in ws.iter_rows(min_row=1, max_col=6, values_only=True):
    print(row)
