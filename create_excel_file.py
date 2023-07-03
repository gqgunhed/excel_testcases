#!/usr/bin/env python

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# create some data
ws.title = "My New TEST Title"
ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
wb.create_sheet("Another sheet")

# see what workbook sheets we have
print(wb.sheetnames)

# write to file
wb.save('test_newly_created.xlsx')
